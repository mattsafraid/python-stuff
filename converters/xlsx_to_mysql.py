#! python3
# xlsx_to_mysql.py - reads from a xlsx file and generates a 
# MySQL database.
# By Matheus Safraid

# REQUIRES: pip install openpyxl
#           pip install lxml

# ASSUMES:  i)  The script is in the same folder as the xlsx file;
#           ii) The xlsx file has a header on line 1. 

# PARAMETERS: <file_name.xlsx>, <db_file.db3>, <max_cols>, <max_lines>

# Why lxml? https://openpyxl.readthedocs.io/en/stable/optimized.html

'''Sample XLSX file: https://data.europa.eu/euodp/en/data/dataset/openfoodtox-efsa-s-chemical-hazards-database'''

from sys import argv
from time import time
from openpyxl import load_workbook
from mysql.connector import connect 
from mysql.connector import Error as myError
from timotheos import beautify
from timotheos import my_create_table_sql, my_insert_sql
from timotheos import xls_fix_header, xls_row_types_mysql

usage = '''{0}
USAGE: python3 {1} <file.xlsx> [<max_cols>] [<max_lines>]
{0}'''.format("="*79, argv[0])

if len(argv) < 2:
    print(usage)
    exit()

########## PARAMETERS ##########
xlsx_file = argv[1]
max_cols  = argv[2] if len(argv) > 3 else 0
max_lines = argv[3] if len(argv) > 4 else 0

insert_buffer_len = 199
col_separator   = "|"
db_name = beautify( xlsx_file.split('.')[0] )

myconfig = {
    'user'      : 'root'
,   'password'  : 'testes'
,   'host'      : '0.0.0.0'
,   'port'      : '3357'
}
########## ENJOY ##########

'''Main'''

"""Create database if not exists"""
try:
    myconn = connect(**myconfig)
    mycur  = myconn.cursor()
    mycur.execute("CREATE DATABASE IF NOT EXISTS {};".format(db_name))
except myError:
    print(myError)
finally:
    myconn.close()
    del mycur
    del myconn
    
"""OK now we're talking"""
myconfig['database'] = db_name


try:
    print(':: Connecting to database {}...'.format(myconfig['database']))
    myconn = connect(**myconfig)
    mycur  = myconn.cursor()
    
    """load_workbook:
    Using read_only=True returns a ReadOnlyWorkSheet
    Using data_only=True returns data instead of formulas."""
    print(':: Loading workbook {}...'.format(xlsx_file))
    wb = load_workbook(xlsx_file, read_only=True, data_only=True)
    
    '''Create a table for each worksheet'''
    for ws in wb:
        print( """:: {0} = {{'lines': {1}, """.format(ws.title, ws.max_row),
              end='')
        
        header = xls_fix_header(ws[1])
        types = list()
        '''Comment the following line if you don't want to infer data types !!!'''
        types  = xls_row_types_mysql(ws[2])
        
        ## Create landing table. 
        print("""'CREATE TABLE': """, end='')
        timer = time()
        sql = my_create_table_sql(
            myconfig['database'], 
            beautify(ws.title), 
            header,
            types)
        mycur.execute(sql)
        myconn.commit()
        print( round(time() - timer, 2) , end='')
        
        ## Insert into table, in batches 
        print(""", 'INSERT': """, end='')
        timer = time()
        sql = my_insert_sql(
            myconfig['database'], 
            beautify(ws.title), 
            header
        )
        buffer = list()
        for i, row in enumerate(ws.iter_rows()):
            # Skip first line
            if i == 0: 
                continue
            is_last_row = (i == ws.max_row - 1)

            # Append only non-empty, non-header lines
            if (row[0].value is not None) and (row[0].value != ws['A1'].value):
                buffer.append(
                    tuple([c.value if c.value != '' else None for c in row]))

            # Insert when buffer is full or when reaches the end
            if len(buffer) >= insert_buffer_len or is_last_row:
                mycur.executemany(sql, buffer)
                myconn.commit()
                buffer.clear()
                
        timer = time() - timer
        print( round(timer, 2), end='' )
        print( """, 'insert_buffer_len': """, insert_buffer_len, end='' )
        print( """, 'lines_per_sec': """, ws.max_row / timer, end='' )
        print("""}""")
        
except Exception as e:
    print(e)
    print(buffer)
finally:
    print(':: Cleaning up...')
    myconn.commit()
    myconn.close()
    wb.close()
    print(':: Done.')
