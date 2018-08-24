#! python3
# xlsx_to_sqlite.py - reads from a xlsx file and generates a 
# sqlite3 database.
# By Matheus Safraid

# REQUIRES: pip install openpyxl
#           pip install lxml

# ASSUMES:  i)  The script is in the same folder as the xlsx file;
#           ii) The xlsx file has a header on line 1. 

# PARAMETERS: <file_name.xlsx>, <db_file.db3>

'''Why lxml? https://openpyxl.readthedocs.io/en/stable/optimized.html'''

from sys import argv
from time import time
from sqlite3  import connect 
from openpyxl import load_workbook
from timotheos import beautify
from timotheos import xls_fix_header, xls_row_types_lite
from timotheos import lite_create_table_sql, lite_insert_sql

usage = '''{0}
USAGE: python3 {1} <file.xlsx> [<dbfile.db3>] [<max_cols>] [<max_lines>]
{0}'''.format("="*79, argv[0])

if len(argv) < 2:
    print(usage)
    exit()

########## PARAMETER AREA ##########
insert_buffer_len = 19

xlsx_file = argv[1]
db3_file  = argv[2] if len(argv) > 2 else xlsx_file.split('.')[0] + '.db3'
########## ENJOY ##########

'''Main'''

try:
    print(':: Connecting to db3 file {}...'.format(db3_file))
    conn3 = connect(db3_file)
    cur3  = conn3.cursor()
    
    print(':: Loading workbook {}...'.format(xlsx_file))
    wb = load_workbook(xlsx_file, read_only=True, data_only=True)
    
    '''Create a table for each worksheet'''
    output = dict()
    for ws in wb:
        output['table'] = ws.title
        output['lines'] = ws.max_row
        
        header = xls_fix_header(ws[1])
        types  = xls_row_types_lite(ws[2])
        
        ## Create table
        timer = time()
        sql = lite_create_table_sql(
            beautify(ws.title), 
            header,
            types)
        cur3.execute(sql)
        conn3.commit()
        output['CREATE TABLE'] = time() - timer
        
        ## Insert into table
        timer = time()
        sql = lite_insert_sql(
            beautify(ws.title), 
            header)
        buffer = list()
        for i, row in enumerate(ws.iter_rows()):
            # Skip first line
            if i == 0:
                continue
            is_last_row = (i == ws.max_row - 1)
            
            # Append non empty, non-header lines
            if(row[0].value is not None) and (row[0].value != ws['A1'].value):
                buffer.append(tuple([c.value for c in row]))
        
            # Insert when buffer is full or when reaches the end
            if len(buffer) >= insert_buffer_len or is_last_row:
                cur3.executemany(sql, buffer)
                conn3.commit()
                buffer.clear()
        output['INSERT'] = time() - timer
        
        output['insert_buffer_len'] = insert_buffer_len
        output['lines per sec'] = ws.max_row /  output['INSERT']
        print(output)
except Exception as e:
    print(e)
finally:
    print(':: Cleaning up...')
    conn3.commit()
    conn3.close()
    wb.close()
    print(':: Done.')